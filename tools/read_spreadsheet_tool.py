import os
from config import ALLOWED_READ_DIRS


def _is_allowed(path: str) -> bool:
    real = os.path.realpath(path)
    return any(real.startswith(os.path.realpath(d)) for d in ALLOWED_READ_DIRS)


class ReadSpreadsheetTool:
    name = "read_spreadsheet"
    description = (
        "Lê planilhas CSV ou Excel (.xlsx/.xls). "
        "Input: {'path': 'arquivo.xlsx', 'sheet': 'Planilha1', 'rows': 50}"
    )

    def run(self, input_data: dict) -> str:
        path = input_data.get("path", "")
        sheet = input_data.get("sheet", 0)
        max_rows = min(int(input_data.get("rows", 50)), 200)

        if not path:
            return "Erro: campo 'path' obrigatório."
        if not _is_allowed(path):
            return f"Bloqueado: '{path}' fora das pastas permitidas."
        if not os.path.exists(path):
            return f"Erro: arquivo não encontrado em '{path}'."

        ext = os.path.splitext(path)[1].lower()

        try:
            if ext == ".csv":
                return self._read_csv(path, max_rows)
            elif ext in (".xlsx", ".xls", ".ods"):
                return self._read_excel(path, sheet, max_rows)
            else:
                return f"Formato não suportado: '{ext}'. Use CSV, XLSX ou XLS."
        except ImportError as e:
            return f"Dependência ausente: {e}. Execute: pip install openpyxl pandas"
        except Exception as e:
            return f"Erro ao ler planilha: {str(e)}"

    def _read_csv(self, path: str, max_rows: int) -> str:
        import csv
        rows = []
        with open(path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.reader(f)
            for i, row in enumerate(reader):
                if i >= max_rows + 1:
                    rows.append([f"... ({i - 1} linhas no total, exibindo {max_rows})"])
                    break
                rows.append(row)
        if not rows:
            return "Planilha CSV vazia."
        return self._format_table(rows, os.path.basename(path))

    def _read_excel(self, path: str, sheet, max_rows: int) -> str:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheet_names = wb.sheetnames

        if isinstance(sheet, int):
            ws = wb.worksheets[sheet] if sheet < len(wb.worksheets) else wb.active
        else:
            ws = wb[sheet] if sheet in sheet_names else wb.active

        rows = []
        total = 0
        for row in ws.iter_rows(values_only=True):
            if total >= max_rows:
                rows.append((f"... ({ws.max_row} linhas no total, exibindo {max_rows})",))
                break
            rows.append(row)
            total += 1

        wb.close()
        if not rows:
            return "Planilha Excel vazia."

        info = f"Arquivo: {os.path.basename(path)} | Aba: {ws.title} | Abas disponíveis: {', '.join(sheet_names)}\n"
        return info + self._format_table(rows, ws.title)

    def _format_table(self, rows: list, name: str) -> str:
        if not rows:
            return "Sem dados."

        str_rows = [[str(c) if c is not None else "" for c in row] for row in rows]
        if not str_rows:
            return "Sem dados."

        col_count = max(len(r) for r in str_rows)
        col_widths = [0] * col_count
        for row in str_rows:
            for i, cell in enumerate(row):
                col_widths[i] = max(col_widths[i], min(len(cell), 30))

        lines = []
        for idx, row in enumerate(str_rows):
            cells = []
            for i in range(col_count):
                val = row[i] if i < len(row) else ""
                val = val[:30] + "…" if len(val) > 30 else val
                cells.append(val.ljust(col_widths[i]))
            lines.append(" | ".join(cells))
            if idx == 0:
                lines.append("-" * (sum(col_widths) + 3 * (col_count - 1)))

        return f"=== {name} ===\n" + "\n".join(lines)
